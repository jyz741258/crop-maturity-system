import os
import json
import csv
from datetime import datetime
from io import BytesIO
import base64

class ExportUtils:
    @staticmethod
    def export_to_csv(data, filename=None):
        """导出数据到CSV文件"""
        if not filename:
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            filename = f'crop_analysis_{timestamp}.csv'
        
        filepath = os.path.join('exports', filename)
        os.makedirs('exports', exist_ok=True)
        
        with open(filepath, 'w', newline='', encoding='utf-8-sig') as csvfile:
            fieldnames = ['检测ID', '作物类型', '成熟度', '置信度(%)', '绿色占比(%)', 
                         '红色占比(%)', '颜色方差', '纹理得分', '形态得分', '品质评分']
            writer = csv.DictWriter(csvfile, fieldnames=fieldnames)
            
            writer.writeheader()
            for item in data:
                writer.writerow({
                    '检测ID': item.get('id', ''),
                    '作物类型': item.get('crop_type', ''),
                    '成熟度': item.get('maturity', ''),
                    '置信度(%)': item.get('confidence', ''),
                    '绿色占比(%)': item.get('green_ratio', ''),
                    '红色占比(%)': item.get('red_ratio', ''),
                    '颜色方差': item.get('color_variance', ''),
                    '纹理得分': item.get('texture_score', ''),
                    '形态得分': item.get('shape_score', ''),
                    '品质评分': item.get('quality_score', '')
                })
        
        return filepath

    @staticmethod
    def export_to_excel(data, filename=None):
        """导出数据到Excel文件"""
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            from openpyxl.chart import PieChart, BarChart, Reference
            from openpyxl.chart.label import DataLabelList
            
            if not filename:
                timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
                filename = f'crop_analysis_{timestamp}.xlsx'
            
            filepath = os.path.join('exports', filename)
            os.makedirs('exports', exist_ok=True)
            
            wb = openpyxl.Workbook()
            
            ws = wb.active
            ws.title = '检测结果'
            
            headers = ['检测ID', '作物类型', '成熟度', '置信度(%)', '绿色占比(%)', 
                       '红色占比(%)', '颜色方差', '纹理得分', '形态得分', '品质评分']
            
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = Font(bold=True, color='FFFFFF')
                cell.fill = PatternFill(start_color='2E7D32', end_color='2E7D32', fill_type='solid')
                cell.alignment = Alignment(horizontal='center', vertical='center')
            
            for row, item in enumerate(data, 2):
                ws.cell(row=row, column=1, value=item.get('id', ''))
                ws.cell(row=row, column=2, value=item.get('crop_type', ''))
                ws.cell(row=row, column=3, value=item.get('maturity', ''))
                ws.cell(row=row, column=4, value=item.get('confidence', ''))
                ws.cell(row=row, column=5, value=item.get('green_ratio', ''))
                ws.cell(row=row, column=6, value=item.get('red_ratio', ''))
                ws.cell(row=row, column=7, value=item.get('color_variance', ''))
                ws.cell(row=row, column=8, value=item.get('texture_score', ''))
                ws.cell(row=row, column=9, value=item.get('shape_score', ''))
                ws.cell(row=row, column=10, value=item.get('quality_score', ''))
            
            for col in range(1, len(headers) + 1):
                ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 15
            
            ws2 = wb.create_sheet(title='统计分析')
            ws2.cell(row=1, column=1, value='统计项')
            ws2.cell(row=1, column=2, value='数值')
            
            maturity_counts = {}
            total_items = len(data)
            avg_confidence = sum(item.get('confidence', 0) for item in data) / total_items if total_items > 0 else 0
            avg_quality = sum(item.get('quality_score', 0) for item in data) / total_items if total_items > 0 else 0
            
            for item in data:
                maturity = item.get('maturity', '未知')
                maturity_counts[maturity] = maturity_counts.get(maturity, 0) + 1
            
            stats = [
                ('检测总数', total_items),
                ('平均置信度', f'{avg_confidence:.1f}%'),
                ('平均品质评分', f'{avg_quality:.1f}')
            ]
            
            for row, (stat, value) in enumerate(stats, 2):
                ws2.cell(row=row, column=1, value=stat)
                ws2.cell(row=row, column=2, value=value)
            
            ws2.cell(row=6, column=1, value='成熟度分布')
            for row, (maturity, count) in enumerate(maturity_counts.items(), 7):
                ws2.cell(row=row, column=1, value=maturity)
                ws2.cell(row=row, column=2, value=count)
            
            pie = PieChart()
            labels = Reference(ws2, min_col=1, min_row=7, max_row=7 + len(maturity_counts) - 1)
            data = Reference(ws2, min_col=2, min_row=7, max_row=7 + len(maturity_counts) - 1)
            pie.add_data(data, titles_from_data=False)
            pie.set_categories(labels)
            pie.title = '成熟度分布'
            pie.dataLabels = DataLabelList()
            pie.dataLabels.showVal = True
            pie.dataLabels.showPercent = True
            ws2.add_chart(pie, 'D2')
            
            wb.save(filepath)
            return filepath
            
        except ImportError:
            return ExportUtils.export_to_csv(data, filename.replace('.xlsx', '.csv'))
        except Exception as e:
            print(f"Excel导出错误: {e}")
            return None

    @staticmethod
    def export_to_pdf(data, image_base64=None, filename=None):
        """导出数据到PDF文件"""
        try:
            from fpdf import FPDF
            
            if not filename:
                timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
                filename = f'crop_analysis_{timestamp}.pdf'
            
            filepath = os.path.join('exports', filename)
            os.makedirs('exports', exist_ok=True)
            
            pdf = FPDF()
            pdf.add_page()
            pdf.set_font('Arial', 'B', 16)
            pdf.cell(0, 20, '作物成熟度检测报告', 0, 1, 'C')
            
            pdf.set_font('Arial', '', 12)
            pdf.cell(0, 10, f'生成时间: {datetime.now().strftime("%Y年%m月%d日 %H:%M:%S")}', 0, 1)
            pdf.cell(0, 10, f'检测数量: {len(data)}', 0, 1)
            pdf.ln(10)
            
            if image_base64:
                try:
                    image_data = base64.b64decode(image_base64.split(',')[1] if ',' in image_base64 else image_base64)
                    image_path = 'temp_export_image.jpg'
                    with open(image_path, 'wb') as f:
                        f.write(image_data)
                    pdf.image(image_path, x=10, y=pdf.get_y(), w=100)
                    os.remove(image_path)
                    pdf.ln(60)
                except:
                    pass
            
            pdf.set_font('Arial', 'B', 12)
            pdf.cell(10, 10, 'ID')
            pdf.cell(25, 10, '成熟度')
            pdf.cell(20, 10, '置信度')
            pdf.cell(20, 10, '品质评分')
            pdf.ln(5)
            
            pdf.set_font('Arial', '', 10)
            for item in data[:20]:
                pdf.cell(10, 8, item.get('id', '')[-4:])
                pdf.cell(25, 8, item.get('maturity', ''))
                pdf.cell(20, 8, f"{item.get('confidence', 0):.1f}%")
                pdf.cell(20, 8, f"{item.get('quality_score', 0):.1f}")
                pdf.ln(5)
            
            maturity_counts = {}
            for item in data:
                maturity = item.get('maturity', '未知')
                maturity_counts[maturity] = maturity_counts.get(maturity, 0) + 1
            
            pdf.ln(10)
            pdf.set_font('Arial', 'B', 12)
            pdf.cell(0, 10, '成熟度分布统计', 0, 1)
            pdf.set_font('Arial', '', 10)
            
            total = len(data)
            for maturity, count in maturity_counts.items():
                percentage = (count / total) * 100 if total > 0 else 0
                pdf.cell(0, 8, f"{maturity}: {count}株 ({percentage:.1f}%)", 0, 1)
            
            pdf.output(filepath)
            return filepath
            
        except ImportError:
            return ExportUtils.export_to_csv(data, filename.replace('.pdf', '.csv'))
        except Exception as e:
            print(f"PDF导出错误: {e}")
            return None

    @staticmethod
    def export_chart_image(chart_data, filename=None):
        """导出图表图片"""
        try:
            import matplotlib
            matplotlib.use('Agg')
            import matplotlib.pyplot as plt
            import seaborn as sns
            
            if not filename:
                timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
                filename = f'chart_{timestamp}.png'
            
            filepath = os.path.join('exports', filename)
            os.makedirs('exports', exist_ok=True)
            
            fig, axes = plt.subplots(1, 2, figsize=(12, 5))
            
            maturity_counts = chart_data.get('maturity_counts', {})
            if maturity_counts:
                axes[0].pie(maturity_counts.values(), labels=maturity_counts.keys(), 
                          autopct='%1.1f%%', colors=['#8bc34a', '#ffc107', '#4caf50', '#e53935'])
                axes[0].set_title('成熟度分布')
            
            trend_data = chart_data.get('trend_data', {})
            if trend_data.get('labels') and trend_data.get('values'):
                axes[1].plot(trend_data['labels'], trend_data['values'], marker='o', color='#2e7d32')
                axes[1].set_title('检测趋势')
                axes[1].tick_params(axis='x', rotation=45)
            
            plt.tight_layout()
            plt.savefig(filepath, dpi=150, bbox_inches='tight')
            plt.close()
            
            return filepath
            
        except ImportError:
            return None
        except Exception as e:
            print(f"图表导出错误: {e}")
            return None

    @staticmethod
    def generate_report(data, format='csv'):
        """生成报告"""
        formats = {
            'csv': ExportUtils.export_to_csv,
            'excel': ExportUtils.export_to_excel,
            'pdf': ExportUtils.export_to_pdf
        }
        
        if format in formats:
            return formats[format](data)
        return None

    @staticmethod
    def get_export_formats():
        """获取支持的导出格式"""
        return ['csv', 'excel', 'pdf', 'image']

    @staticmethod
    def generate_summary(data):
        """生成数据摘要"""
        if not data:
            return {}
        
        total = len(data)
        maturity_counts = {}
        total_confidence = 0
        total_quality = 0
        total_green_ratio = 0
        
        for item in data:
            maturity = item.get('maturity', '未知')
            maturity_counts[maturity] = maturity_counts.get(maturity, 0) + 1
            total_confidence += item.get('confidence', 0)
            total_quality += item.get('quality_score', 0)
            total_green_ratio += item.get('green_ratio', 0)
        
        return {
            'total_count': total,
            'maturity_distribution': maturity_counts,
            'avg_confidence': round(total_confidence / total, 1) if total > 0 else 0,
            'avg_quality': round(total_quality / total, 1) if total > 0 else 0,
            'avg_green_ratio': round(total_green_ratio / total, 1) if total > 0 else 0,
            'mature_rate': round(maturity_counts.get('成熟期', 0) / total * 100, 1) if total > 0 else 0
        }

    @staticmethod
    def export_as_json(data, filename=None):
        """导出数据为JSON格式"""
        if not filename:
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            filename = f'crop_analysis_{timestamp}.json'
        
        filepath = os.path.join('exports', filename)
        os.makedirs('exports', exist_ok=True)
        
        with open(filepath, 'w', encoding='utf-8') as f:
            json.dump(data, f, ensure_ascii=False, indent=2)
        
        return filepath